#! /usr/bin/python
# -*- coding: utf-8 -*-
#===============================================================================#
#      Feronia - Builds biodiversity databases from species checklists          #
#                   (C) 2016 by Mauro J. Cavalcanti                             #
#                         <maurobio@gmail.com>                                  #
#                                                                               #
#  This program is free software; you can redistribute it and/or modify         #
#  it under the terms of the GNU General Public License as published by         #
#  the Free Software Foundation; either version 3 of the License, or            #
#  (at your option) any later version.                                          #
#                                                                               #
#  This program is distributed in the hope that it will be useful,              #
#  but WITHOUT ANY WARRANTY; without even the implied warranty of               #
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the                # 
#  GNU General Public License for more details.                                 #
#                                                                               #
#  You should have received a copy of the GNU General Public License            #
#  along with this program. If not, see <http://www.gnu.org/licenses/>.         #
#                                                                               #       
#  Requirements:                                                                #
#    Python 2.7+ (www.python.org)                                               #
#    PyQt 4.8+ (www.riverbankcomputing.com/software/pyqt)                       #
#    formlayout 1.0+ (code.google.com/p/formlayout)                             #
#    lxml 3.6+ (lxml.de)                                                        #
#    openpyxl 2.0+ (openpyxl.readthedocs.org/en/2.0)                            #
#    ezodf 0.2+ (pythonhosted.org/ezodf)                                        #
#    xlrd (www.python-excel.org)                                                #
#    MySQLdb (sourceforge.net/projects/mysql-python)                            #
#    mysql.connector (dev.mysql.com/downloads/connector/python/2.1.html)        #
#    psycopg2 (initd.org/psycopg)                                               # 
#    fdb (www.firebirdsql.org/en/devel-python-driver)                           #
#    pygbif 0.1+ (github.com/sckott/pygbif)                                     #
#    BioPython (biopython.org/wiki/Main_Page)                                   #
#    wikipedia (github.com/goldsmith/Wikipedia)                                 #
#                                                                               #
#  REVISION HISTORY:                                                            #
#    Version 1.00, 17th May 16 - Initial release                                #
#===============================================================================#                                                 

import sys
import os
import os.path
import time
import platform
import warnings
import unicodedata
import simplejson
import urllib
import lxml.etree as ET
from os.path import basename
from PyQt4 import QtCore, QtGui
from formlayout import fedit
import resources

def encode_for_xml(unicode_data, encoding="ascii"):
    return unicode_data.encode(encoding, "xmlcharrefreplace")

def iif(boolVar, ifTrue, ifFalse):
    if boolVar:
        return ifTrue
    else:
        return ifFalse

def is_ascii(s):
    return all(ord(c) < 128 for c in s)

def unicode_to_ascii(str):
    return unicodedata.normalize("NFKD", unicode(str)).encode("ascii","ignore")

#--- Disable all warnings
warnings.filterwarnings("ignore")

__version__ = "1.0.2"

class MainWindow(QtGui.QMainWindow):

    def __init__(self):
        super(MainWindow, self).__init__()
        self.initUI()

    def initUI(self):
        self.data = []
        self.db = None
        self.adapter = ""
        self.filename = ""
        
        self.openAction = QtGui.QAction("&Open", self)
        self.openAction.setShortcut("Ctrl+O")
        self.openAction.setStatusTip("Open a file")
        self.openAction.triggered.connect(self.openFile)
        
        self.connectAction = QtGui.QAction("&Connect", self)
        self.connectAction.setStatusTip("Connect to a database")
        self.connectAction.triggered.connect(self.connectDb)

        self.closeAction = QtGui.QAction("&Quit", self)
        self.closeAction.setShortcut("Ctrl+Q")
        self.closeAction.setStatusTip("Close application")
        self.closeAction.triggered.connect(self.close)
        
        self.helpAction = QtGui.QAction("&About", self)
        self.helpAction.setStatusTip("Display information")
        self.helpAction.triggered.connect(self.about)
        
        self.bibliographyAction = QtGui.QAction("&Bibliography", self)
        self.bibliographyAction.setStatusTip("Bibliography")
        self.bibliographyAction.triggered.connect(self.bibliography)
        
        self.conservationAction = QtGui.QAction("&Conservation", self)
        self.conservationAction.setStatusTip("Conservation Status")
        self.conservationAction.triggered.connect(self.conservation)
        
        self.distributionAction = QtGui.QAction("&Distribution", self)
        self.distributionAction.setStatusTip("Geographic Distribution")
        self.distributionAction.triggered.connect(self.distribution)
        
        self.genomeAction = QtGui.QAction("&Genome", self)
        self.genomeAction.setStatusTip("Genome")
        self.genomeAction.triggered.connect(self.genome)
        
        self.habitatsAction = QtGui.QAction("&Habitats", self)
        self.habitatsAction.setStatusTip("Habitats")
        self.habitatsAction.triggered.connect(self.habitats)
        
        self.notesAction = QtGui.QAction("&Notes", self)
        self.notesAction.setStatusTip("Structured Notes")
        self.notesAction.triggered.connect(self.notes)
        
        self.synonymsAction = QtGui.QAction("&Synonyms", self)
        self.synonymsAction.setStatusTip("Synonyms")
        self.synonymsAction.triggered.connect(self.synonyms)
        
        ##self.taxaAction = QtGui.QAction("&Taxa", self)
        ##self.taxaAction.setStatusTip("Taxa")
        ##self.taxaAction.triggered.connect(self.taxa)
        
        self.commonnamesAction = QtGui.QAction("&Vernacular", self)
        self.commonnamesAction.setStatusTip("Vernacular Names")
        self.commonnamesAction.triggered.connect(self.commonnames)

        menubar = self.menuBar()
        fileMenu = menubar.addMenu("&File")
        fileMenu.addAction(self.openAction)
        fileMenu.addAction(self.connectAction)
        fileMenu.addSeparator()
        fileMenu.addAction(self.closeAction)
        
        dataMenu = menubar.addMenu("&Data")
        dataMenu.addAction(self.bibliographyAction)
        dataMenu.addAction(self.conservationAction)
        dataMenu.addAction(self.distributionAction)
        dataMenu.addAction(self.genomeAction)
        dataMenu.addAction(self.habitatsAction)
        dataMenu.addAction(self.notesAction)
        dataMenu.addAction(self.synonymsAction)
        ##dataMenu.addAction(self.taxaAction)
        dataMenu.addAction(self.commonnamesAction)
        
        helpMenu = menubar.addMenu("&Help")
        helpMenu.addAction(self.helpAction)

        statusbar = self.statusBar()
        statusbar.setSizeGripEnabled(True)
        statusbar.showMessage("Ready")
        
        self.text = QtGui.QTextEdit(self)
        self.text.setReadOnly(True)
        self.text.setHtml(
            """Builds biodiversity databases from species checklists. 
            <br>&copy; 2016 Mauro J. Cavalcanti. 
            <br>Ecoinformatics Studio, Rio de Janeiro, Brazil. 
            <br>E-mail: maurobio@gmail.com""")
        
        self.setCentralWidget(self.text)
        self.setGeometry(100,100,650,350)
        self.setWindowTitle("Feronia")
        self.setWindowIcon(QtGui.QIcon(":/icon.png"))
        self.updateUI()
        self.show()
        
    def updateUI(self):
        self.connectAction.setEnabled(len(self.data) > 0)
        enable = self.db is not None
        self.bibliographyAction.setEnabled(enable)
        self.conservationAction.setEnabled(enable)
        self.distributionAction.setEnabled(enable)
        self.genomeAction.setEnabled(enable)
        self.habitatsAction.setEnabled(enable)
        self.notesAction.setEnabled(enable)
        self.synonymsAction.setEnabled(enable)
        ##self.taxaAction.setEnabled(enable)
        self.commonnamesAction.setEnabled(enable)

    def readData(self, filename):
        file_ext = filename[-3:]
        self.data = []
        if file_ext == "xls":
            import xlrd
            wb = xlrd.open_workbook(filename)
            sh1 = wb.sheet_by_index(0)
            for rownum in range(sh1.nrows): 
                self.data += [sh1.row_values(rownum)]
        elif file_ext == "csv":
            import csv
            reader = csv.reader(open(filename, "rb"))
            for row in reader:
                self.data += [row]
        elif file_ext == "lsx":
            from openpyxl.reader.excel import load_workbook
            wb = load_workbook(filename=filename, use_iterators = True)
            sheet = wb.get_active_sheet()
            for row in sheet.iter_rows():
                data_row = []
                for cell in row:
                    data_row += [cell.value]
                self.data += [data_row]
        elif file_ext == "ods":
            from ezodf import opendoc
            wb = opendoc(filename)
            s = wb.sheets[0]
            for row in range(s.nrows()-1):
                values = []
                for col in range(s.ncols()-1):
                    values += [s[(row,col)].value]
                self.data += [values]
        self.text.append("<br><b>Read " + str(len(self.data)-1) + " records from file '" + basename(filename) + "'</b>")
        del self.data[0]
    
    def connectDb(self):
        options = [("Adapter:",
                    [0, "MySQL",
                        "MariaDB",
                        "PostgreSQL",
                        "Firebird",
                        "SQLite"]),
                    ("Username:", ""),
                    ("Password:", ""),
                    ("Database:", "")
                    ]
        
        self.db = None
        while True:
            result = fedit(options,
                        title="Database Adapter",
                        icon=QtGui.QIcon(":/icon.png"),
                        parent=self)
            if result is None: break
        
            adapter = result[0]
            user = result[1]
            pwd = result[2]
            db_name = result[3]
            self.adapter = options[0][1][adapter+1]
            try:
                if adapter == 0:
                    import MySQLdb
                    self.db = MySQLdb.connect(host="localhost", user=user, passwd=pwd, db=db_name)
                elif adapter == 1:
                    import mysql.connector as mariadb
                    self.db = mariadb.connect(host="localhost", user=user, passwd=pwd, db=db_name)
                elif adapter == 2:
                    import psycopg2
                    self.db = psycopg2.connect(host="localhost", user=user, passwd=pwd, db=db_name)
                elif adapter == 3:
                    import fdb
                    self.db = fdb.connect(host="localhost", user=user, passwd=pwd, db=db_name)
                elif adapter == 4:
                    import sqlite3
                    self.db = sqlite3.connect(db_name)
                    self.db.text_factory = str
                #self.text.append("<br><b>Database '" + db_name + "' connected using " + options[0][1][adapter+1] + " adapter</b>")
                self.text.append("<br><b>Database '" + db_name + "' connected using " + self.adapter + " adapter</b>")
                break
            except Exception, e:
                QtGui.QMessageBox.critical(self, "Error", str(e[1]))
        self.updateUI()
        
    def createTables(self):
        with self.db:
            cursor = self.db.cursor()
            
            #--- Bibliography table (from CoL / EOL)
            ##cursor.execute("DROP TABLE IF EXISTS bibliography")
            cursor.execute("CREATE TABLE IF NOT EXISTS bibliography(B_NO INT PRIMARY KEY, B_TYPE VARCHAR(20), B_AUTHOR VARCHAR(128), \
                B_YEAR INT, B_SEQUENCE CHAR(1), B_TITLE VARCHAR(254), B_DETAIL VARCHAR(512))")
            
            #--- Conservation status table (from IUCN)
            ##cursor.execute("DROP TABLE IF EXISTS status")
            if self.adapter == "SQLite":
                cursor.execute("CREATE TABLE IF NOT EXISTS status(ID INTEGER PRIMARY KEY AUTOINCREMENT, T_NO INT, C_STATUS VARCHAR(22), C_TREND VARCHAR(12), B_NO INT)")
            else:    
                cursor.execute("CREATE TABLE IF NOT EXISTS status(ID INT PRIMARY KEY AUTO_INCREMENT, T_NO INT, C_STATUS VARCHAR(22), C_TREND VARCHAR(12), B_NO INT)")
            
            #--- Geographic distribution table (from GBIF)
            ##cursor.execute("DROP TABLE IF EXISTS distribution")
            if self.adapter == "SQLite":
                cursor.execute("CREATE TABLE IF NOT EXISTS distribution(ID INTEGER PRIMARY KEY AUTOINCREMENT, T_NO INT, P_CODE VARCHAR(50), P_CONTINENT VARCHAR(20), \
                    P_REGION VARCHAR(30), P_COUNTRY VARCHAR(30), P_STATE VARCHAR(30), P_COUNTY VARCHAR(30), \
                    P_LOCALITY VARCHAR(254), P_LATITUDE FLOAT, P_LONGITUDE FLOAT, P_I_STATUS VARCHAR(10), B_NO INT)")
            else:
                cursor.execute("CREATE TABLE IF NOT EXISTS distribution(ID INT PRIMARY KEY AUTO_INCREMENT, T_NO INT, P_CODE VARCHAR(50), P_CONTINENT VARCHAR(20), \
                    P_REGION VARCHAR(30), P_COUNTRY VARCHAR(30), P_STATE VARCHAR(30), P_COUNTY VARCHAR(30), \
                    P_LOCALITY VARCHAR(254), P_LATITUDE FLOAT, P_LONGITUDE FLOAT, P_I_STATUS VARCHAR(10), B_NO INT)")

            #--- Genome table (from NCBI)
            ##cursor.execute("DROP TABLE IF EXISTS genome")
            if self.adapter == "SQLite":
                cursor.execute("CREATE TABLE IF NOT EXISTS genome(ID INTEGER PRIMARY KEY AUTOINCREMENT, T_NO INT, G_TAXID INT, G_SEQ_ID INT, G_SEQ_TYPE VARCHAR(12), \
                    G_DESCRIPTION VARCHAR(254), G_SEQUENCE TEXT, B_NO INT)")
            else:
                cursor.execute("CREATE TABLE IF NOT EXISTS genome(ID INT PRIMARY KEY AUTO_INCREMENT, T_NO INT, G_TAXID INT, G_SEQ_ID INT, G_SEQ_TYPE VARCHAR(12), \
                    G_DESCRIPTION VARCHAR(254), G_SEQUENCE TEXT, B_NO INT)")
            
            #--- Habitats table (from IUCN)
            ##cursor.execute("DROP TABLE IF EXISTS habitats")
            if self.adapter == "SQLite":
                cursor.execute("CREATE TABLE IF NOT EXISTS habitats(ID INTEGER PRIMARY KEY AUTOINCREMENT, T_NO INT, H_PLACE VARCHAR(30), H_HABITAT VARCHAR(78), B_NO INT)")
            else:
                cursor.execute("CREATE TABLE IF NOT EXISTS habitats(ID INT PRIMARY KEY AUTO_INCREMENT, T_NO INT, H_PLACE VARCHAR(30), H_HABITAT VARCHAR(78), B_NO INT)")
            
            #--- Metadata table (from user)
            ##cursor.execute("DROP TABLE IF EXISTS metadata")
            cursor.execute("CREATE TABLE IF NOT EXISTS metadata(M_ID INT PRIMARY KEY, M_ACRONYM VARCHAR(50), M_TITLE VARCHAR(128), \
                M_DESCRIPTION VARCHAR(255), M_SCOPE VARCHAR(20), M_ENVIRONMENT VARCHAR(20), \
                M_COVERAGE VARCHAR(255), M_AUTHOR VARCHAR(255), M_VERSION VARCHAR(128), \
                M_DATE DATE, M_PUBLISHER VARCHAR(128), M_URL VARCHAR(128), M_LOGO VARCHAR(128), \
                M_BANNER VARCHAR(128))")
            
            #--- Notes table (from Wikipedia)
            ##cursor.execute("DROP TABLE IF EXISTS notes")
            if self.adapter == "SQLite":
                cursor.execute("CREATE TABLE IF NOT EXISTS notes(ID INTEGER PRIMARY KEY AUTOINCREMENT, T_NO INT, N_NOTE VARCHAR(255), B_NO INT)")
            else:
                cursor.execute("CREATE TABLE IF NOT EXISTS notes(ID INT PRIMARY KEY AUTO_INCREMENT, T_NO INT, N_NOTE VARCHAR(255), B_NO INT)")
            
            #--- Literature pointers
            ##cursor.execute("DROP TABLE IF EXISTS pointers")
            if self.adapter == "SQLite":
                cursor.execute("CREATE TABLE IF NOT EXISTS pointers(ID INTEGER PRIMARY KEY AUTOINCREMENT, T_NO INT, L_TYPE VARCHAR(20), B_NO INT)")
            else:
                cursor.execute("CREATE TABLE IF NOT EXISTS pointers(ID INT PRIMARY KEY AUTO_INCREMENT, T_NO INT, L_TYPE VARCHAR(20), B_NO INT)")
            
            #--- Media resources table (from EOL / Wikipedia)
            ##cursor.execute("DROP TABLE IF EXISTS resources")
            if self.adapter == "SQLite":
                cursor.execute("CREATE TABLE IF NOT EXISTS resources(ID INTEGER PRIMARY KEY AUTOINCREMENT, T_NO INT, R_TYPE VARCHAR(20), \
                    R_RESOURCE VARCHAR(128), R_CAPTION VARCHAR(255), B_NO INT)")
            else:
                cursor.execute("CREATE TABLE IF NOT EXISTS resources(ID INT PRIMARY KEY AUTO_INCREMENT, T_NO INT, R_TYPE VARCHAR(20), \
                    R_RESOURCE VARCHAR(128), R_CAPTION VARCHAR(255), B_NO INT)")
            
            #--- Synonyms table (from CoL)
            ##cursor.execute("DROP TABLE IF EXISTS synonyms")
            if self.adapter == "SQLite":
                cursor.execute("CREATE TABLE IF NOT EXISTS synonyms(ID INTEGER PRIMARY KEY AUTOINCREMENT, T_NO INT, S_STATUS VARCHAR(22), S_GENUS VARCHAR(30), \
                    S_G_AUTHOR VARCHAR(40), S_SUBGENUS VARCHAR(30), S_SPECIES VARCHAR(30), S_S_AUTHOR VARCHAR(50), \
                    S_RANK VARCHAR(7), S_SUBSP VARCHAR(30), S_SP_AUTHOR VARCHAR(40), B_NO INT)")
            else:
                cursor.execute("CREATE TABLE IF NOT EXISTS synonyms(ID INT PRIMARY KEY AUTO_INCREMENT, T_NO INT, S_STATUS VARCHAR(22), S_GENUS VARCHAR(30), \
                    S_G_AUTHOR VARCHAR(40), S_SUBGENUS VARCHAR(30), S_SPECIES VARCHAR(30), S_S_AUTHOR VARCHAR(50), \
                    S_RANK VARCHAR(7), S_SUBSP VARCHAR(30), S_SP_AUTHOR VARCHAR(40), B_NO INT)")
            
            #--- Common names table (from CoL)
            ##cursor.execute("DROP TABLE IF EXISTS commonnames")
            if self.adapter == "SQLite":
                cursor.execute("CREATE TABLE IF NOT EXISTS commonnames(ID INTEGER PRIMARY KEY AUTOINCREMENT, T_NO INT, V_NAME VARCHAR(40), V_COUNTRY VARCHAR(30), \
                    V_LANGUAGE VARCHAR(30), B_NO INT)")
            else:
                cursor.execute("CREATE TABLE IF NOT EXISTS commonnames(ID INT PRIMARY KEY AUTO_INCREMENT, T_NO INT, V_NAME VARCHAR(40), V_COUNTRY VARCHAR(30), \
                    V_LANGUAGE VARCHAR(30), B_NO INT)")
            
            #--- Taxa table (from user-provided species checklist)
            ##cursor.execute("DROP TABLE IF EXISTS taxa")
            cursor.execute("CREATE TABLE IF NOT EXISTS taxa(T_NO INT PRIMARY KEY, T_STATUS VARCHAR(22), T_GENUS VARCHAR(30), \
                T_G_AUTHOR VARCHAR(40), T_SUBGENUS VARCHAR(30), T_SPECIES VARCHAR(30), T_S_AUTHOR VARCHAR(50), \
                T_RANK VARCHAR(7), T_SUBSP VARCHAR(30), T_SP_AUTHOR VARCHAR(40), B_NO INT)")
            
            #--- Higher taxa table (from user-provided species checklist)
            ##cursor.execute("DROP TABLE IF EXISTS highertaxa")
            cursor.execute("CREATE TABLE IF NOT EXISTS highertaxa(T_NO INT PRIMARY KEY, T_KINGDOM VARCHAR(50), T_PHYLUM VARCHAR(50), \
                T_SUBPHYLUM VARCHAR(50), T_CLASS VARCHAR(50), T_SUBCLASS VARCHAR(50), T_ORDER VARCHAR(50), \
                T_SUBORDER VARCHAR(50), T_FAMILY VARCHAR(50), T_SUPERFAMILY VARCHAR(50), T_SUBFAMILY VARCHAR(50), \
                T_TRIBE VARCHAR(50))")
            cursor.close()
        
    def bibliography(self):
        with self.db:
            cursor = self.db.cursor()
            cursor.execute("SELECT COUNT(*) from bibliography")
            n = int(cursor.fetchone()[0])
            if n > 0:
                reply = QtGui.QMessageBox.question(self, "Confirmation", 
                        "Delete data?\nThis operation cannot be undone!", 
                        QtGui.QMessageBox.Yes, QtGui.QMessageBox.No)
                if reply == QtGui.QMessageBox.Yes:
                        cursor.execute("DROP TABLE IF EXISTS bibliography")
                else:
                    return
            
        reccount = 0    
        referencelist = []
        self.text.append("<br><b>Fetching bibliographic references from CoL...</b><br>")
        for i in range(len(self.data)):
            #-- For each taxon in databaee
            TaxNo = i + 1
            Genus = self.data[i][5]
            Species = self.data[i][6]
            Subsp = self.data[i][8]
            Name = Genus + ' ' + Species + ' ' + Subsp
    
            #-- Retrieve data from CoL
            result = urllib.urlopen("http://www.catalogueoflife.org/col/webservice?name=" + urllib.quote_plus(Name) + "&response=full").read()
            root = ET.XML(result)
            
            #-- Get a list of references
            references = root.xpath("result/references/reference")
            self.text.append(str(reccount + 1) + " <i>" + Genus + ' ' + Species + ' ' + Subsp + "</i> " + ' - ' + str(len(references)) + " reference(s)")
    
            #-- Loop through the references
            Type = "article"
            Sequence = ''
            for reference in references:
                #-- Get the reference information
                try:
                    Author = reference.xpath("author/text()")[0]
                except IndexError:
                    Author = ""
                try:
                    Year = reference.xpath("year/text()")[0]
                except IndexError:
                    Year = 0
                try:
                    Title = reference.xpath("title/text()")[0]
                except IndexError:
                    Title = ""
                try:
                    Source = reference.xpath("source/text()")[0]
                except IndexError:
                    Source = ""
                item = (Type, Author, Year, Sequence, Title, Source)
                referencelist.append(item)
            
            reccount += 1
            self.text.repaint()
            
            # Sleep for one second to prevent IP blocking from CoL
            time.sleep(1)

        #-- Remove deuplicates from reference list
        uniquelist = set(referencelist)
        referencelist = list(set(uniquelist))

        #-- Add numbers to reference list
        numberedlist = [tuple([index+1] + list(ref)) for index, ref in enumerate(referencelist)]
        referencelist = numberedlist
        refcount = len(referencelist) 

        #-- Insert records into the database 
        try:
            cursor = self.db.cursor()
            if self.adapter == "SQLite":
                cursor.executemany("INSERT INTO bibliography \
                    (B_NO, B_TYPE, B_AUTHOR, B_YEAR, B_SEQUENCE, B_TITLE, B_DETAIL) \
                    VALUES (?, ?, ?, ?, ?, ?, ?)",
                    referencelist)
            else:
                cursor.executemany("INSERT INTO bibliography \
                    (B_NO, B_TYPE, B_AUTHOR, B_YEAR, B_SEQUENCE, B_TITLE, B_DETAIL) \
                    VALUES (%s, %s, %s, %s, %s, %s, %s)",
                    referencelist)
            self.db.commit()
    
        except Exception, e:
            QtGui.QMessageBox.critical(self, "Error", str(e[1]))
            self.db.rollback()

        finally:
            cursor.close()
            
        self.text.append("<br>" + str(reccount) + " record(s)  processed<br>")
    
    def conservation(self):
        with self.db:
            cursor = self.db.cursor()
            cursor.execute("SELECT COUNT(*) from status")
            n = int(cursor.fetchone()[0])
            if n > 0:
                reply = QtGui.QMessageBox.question(self, "Confirmation", 
                        "Delete data?\nThis operation cannot be undone!", 
                        QtGui.QMessageBox.Yes, QtGui.QMessageBox.No)
                if reply == QtGui.QMessageBox.Yes:
                        cursor.execute("DROP TABLE IF EXISTS status")
                else:
                    return
        
        SEARCH_BASE = 'http://apiv3.iucnredlist.org/api/v3/'
        SEARCH_SPECIES = "species/"
        TOKEN = "<YOUR_API_KEY>"
        
        category = {"NE": "Not Evaluated",
                  "DD": "Data Deficient",
                  "LC": "Least Concern",
                  "NT": "Not Threatened",
                  "VU": "Vulnerable",
                  "EN": "Endangered",
                  "CR": "Critically Endangered",
                  "EW": "Extinct in the Wild",
                  "EX": "Extinct"  
                }
                 
        reccount = 0    
        conservation = []
        RefNo = 0
        self.text.append("<br><b>Fetching conservation status from IUCN...</b><br>")
        for i in range(len(self.data)):
            #-- For each taxon in databaee
            TaxNo = i + 1
            Genus = self.data[i][5]
            Species = self.data[i][6]
            Subsp = self.data[i][8]
            Name = Genus + ' ' + Species + iif(len(Subsp) > 0, ' ' + Subsp, "")
            
            #-- Get conservation status from IUCN
            url = SEARCH_BASE + SEARCH_SPECIES + Name + "?token=" + TOKEN
            results = simplejson.load(urllib.urlopen(url))
            result = results['result']
            Trend = "Unknown"
            try:
                Status = category[result[0]['category']]
            except IndexError:
                Status = category["NE"]
            self.text.append(str(reccount + 1) + " <i>" + Genus + ' ' + Species + ' ' + Subsp + "</i> " + " - Status: " + Status)

            item = (TaxNo, Status, Trend, RefNo)
            conservation.append(item)
            
            reccount += 1
            self.text.repaint()
    
            # Sleep for two seconds to prevent IP blocking from IUCN
            time.sleep(2)
                
        #-- Insert records into the database 
        try:
            #-- Insert records into the database 
            cursor = self.db.cursor()
            if self.adapter == "SQLite":
                cursor.executemany("INSERT INTO status \
                    (T_NO, C_STATUS, C_TREND, B_NO) \
                    VALUES (?, ?, ?, ?)",
                    conservation)
            else:
                cursor.executemany("INSERT INTO status \
                    (T_NO, C_STATUS, C_TREND, B_NO) \
                    VALUES (%s, %s, %s, %s)",
                    conservation)
            self.db.commit()
    
        except Exception, e:
            QtGui.QMessageBox.critical(self, "Error", str(e[1]))
            self.db.rollback()

        finally:
            cursor.close()
            
        self.text.append("<br>" + str(reccount) + " record(s)  processed<br>")
    
    def distribution(self):
        from pygbif import species, occurrences
        
        with self.db:
            cursor = self.db.cursor()
            cursor.execute("SELECT COUNT(*) from distribution")
            n = int(cursor.fetchone()[0])
            if n > 0:
                reply = QtGui.QMessageBox.question(self, "Confirmation", 
                        "Delete data?\nThis operation cannot be undone!", 
                        QtGui.QMessageBox.Yes, QtGui.QMessageBox.No)
                if reply == QtGui.QMessageBox.Yes:
                        cursor.execute("DROP TABLE IF EXISTS distribution")
                else:
                    return
        
        reccount = 0
        distribution = []
        RefNo = 0
        self.text.append("<br><b>Fetching distribution data from GBIF...</b><br>")
        for i in range(len(self.data)):
            #-- For each taxon in databaee
            TaxNo = i + 1
            Genus = self.data[i][5]
            Species = self.data[i][6]
            Subsp = self.data[i][8]
            Name = Genus + ' ' + Species + ' ' + Subsp
            
            #-- Get geographic distribution data from GBIF
            key = species.name_backbone(name=Name, rank="species")["usageKey"]
            n = occurrences.count(taxonKey=key, isGeoreferenced=True)
            if n > 300: 
                max = 300
            else:
                max = n
            results = occurrences.search(taxonKey=key, limit=max)
            self.text.append(str(reccount + 1) + " <i>" + Genus + ' ' + Species + ' ' + Subsp + "</i> " + ' - ' + format(n, ',') + " occurrence(s)")
            for x in results["results"]:
                Region = None
                Status = None
        
                try:
                    Code = x['institutionCode'] + '-' + x['catalogNumber']
                except:
                    Code = ""
        
                try:
                    if x['continent'].find('_') != -1:
                        Continent = ' '.join(x['continent'].split('_')).title()
                    else:
                        Continent = x['continent'].capitalize()
                except:
                    Continent = ""
            
                try:
                    Country = x['country'].encode('latin-1', 'replace')
                except:
                    Country = ""
            
                try:
                    State = x['stateProvince'].encode('latin-1', 'replace')
                except:
                    State = ""
            
                try:
                    County = x['county'].encode('latin-1', 'replace')
                except:
                    County = ""
            
                try:
                    Locality = x['locality'].encode('latin-1', 'replace')
                except:
                    Locality = ""
            
                try:
                    Latitude = x['decimalLatitude']
                except:
                    Latitude = 0.0
        
                try:
                    Longitude = x['decimalLongitude']
                except:
                    Longitude = 0.0
            
                item = (TaxNo, Code, Continent, Region, Country, State, County, Locality, Latitude, Longitude, Status, RefNo)
                distribution.append(item)
    
            reccount += 1
            self.text.repaint()
                
            # Sleep for two seconds to prevent IP blocking from GBIF
            time.sleep(1)
                
        #-- Insert records into the database 
        try:
            cursor = self.db.cursor()
            if self.adapter == "SQLite":
                cursor.executemany("INSERT INTO distribution \
                    (T_NO, P_CODE, P_CONTINENT, P_REGION, P_COUNTRY, P_STATE, P_COUNTY, P_LOCALITY, P_LATITUDE, P_LONGITUDE, P_I_STATUS, B_NO) \
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    distribution)
            else:
                cursor.executemany("INSERT INTO distribution \
                    (T_NO, P_CODE, P_CONTINENT, P_REGION, P_COUNTRY, P_STATE, P_COUNTY, P_LOCALITY, P_LATITUDE, P_LONGITUDE, P_I_STATUS, B_NO) \
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    distribution)
            self.db.commit()
    
        except Exception, e:
            QtGui.QMessageBox.critical(self, "Error", str(e[1]))
            db.rollback()
                
        finally:
            cursor.close()
                
        self.text.append("<br>" + str(reccount) + " record(s)  processed<br>")
    
    def genome(self):
        from Bio import SeqIO, Entrez
        
        with self.db:
            cursor = self.db.cursor()
            cursor.execute("SELECT COUNT(*) from genome")
            n = int(cursor.fetchone()[0])
            if n > 0:
                reply = QtGui.QMessageBox.question(self, "Confirmation", 
                        "Delete data?\nThis operation cannot be undone!", 
                        QtGui.QMessageBox.Yes, QtGui.QMessageBox.No)
                if reply == QtGui.QMessageBox.Yes:
                        cursor.execute("DROP TABLE IF EXISTS genome")
                else:
                    return
            
        #-- Save Genbank results to temporary file
        tempfile = "temp.dat"

        #-- Always tell NCBI who you are
        Entrez.email = "<your_email_address>"
        
        reccount = 0
        sequences = []
        RefNo = 0
        self.text.append("<br><b>Fetching genomic data from NCBI...</b><br>")
        for i in range(len(self.data)):
            #-- For each taxon in databaee
            TaxNo = i + 1
            Genus = self.data[i][5]
            Species = self.data[i][6]
            Subsp = self.data[i][8]
            Name = Genus + ' ' + Species + iif(len(Subsp) > 0, ' ' + Subsp, "")
            Taxon = Genus + '+' + Species + iif(len(Subsp) > 0, '+' + Subsp, "")
            
            try:
                #-- Retrieve taxonomy from NCBI
                handle = Entrez.esearch(db="taxonomy",term=Taxon)
                record = Entrez.read(handle)
                try:
                    TaxID = record['IdList'][0]
                    handle = Entrez.esummary(db="taxonomy", id=TaxID)
                    record = Entrez.read(handle)
        
                    #-- Retrieve nucleotide sequences
                    handle = Entrez.esearch(db="nucleotide",term=Taxon)
                    record = Entrez.read(handle)
                    nucl = int(record['RetMax'])
                    for i in range(nucl):
                        #-- Download sequence
                        SeqID = record['IdList'][i]
                        net_handle = Entrez.efetch(db="nucleotide", id=SeqID, rettype="fasta", retmode="text")
                        out_handle = open(tempfile, "w")
                        out_handle.write(net_handle.read())
                        out_handle.close()
                        net_handle.close()
                        seq_record = SeqIO.read(tempfile, "fasta")
                        sequence = (TaxNo, TaxID, SeqID, "nucleotide", seq_record.description, str(seq_record.seq), RefNo)
                        sequences.append(sequence)
                except:
                    nucl = 0
                
                try:
                    #-- Retrieve protein sequences from NCBI
                    handle = Entrez.esearch(db="protein",term=Taxon)
                    record = Entrez.read(handle)
                    prot = int(record['RetMax'])
                    for j in range(prot):
                        #-- Download sequence
                        seqid = record['IdList'][j]
                        net_handle = Entrez.efetch(db="protein", id=SeqID, rettype="fasta", retmode="text")
                        out_handle = open(tempfile, "w")
                        out_handle.write(net_handle.read())
                        out_handle.close()
                        net_handle.close()
                        seq_record = SeqIO.read(tempfile, "fasta")
                        sequence = (TaxNo, TaxID, SeqID, "protein", seq_record.description, str(seq_record.seq), RefNo)
                        sequences.append(sequence)
                except:
                    prot = 0
            
            except:
                continue
            
            self.text.append(str(reccount + 1) + " <i>" + Genus + ' ' + Species + ' ' + Subsp + "</i> " + " - Sequences: " + "nucleotide: " + str(nucl) + ', ' + "protein: " + str(prot))
    
            reccount += 1
            self.text.repaint()
    
            # Sleep for one second to prevent IP blocking from NCBI
            time.sleep(1)
            
        if os.path.exists(tempfile):
            os.remove(tempfile)

        #-- Insert records into the database 
        try:
            cursor = self.db.cursor()
            if self.adapter == "SQLite":
                cursor.executemany("INSERT INTO genome \
                    (T_NO, G_TAXID, G_SEQ_ID, G_SEQ_TYPE, G_DESCRIPTION, G_SEQUENCE, B_NO) \
                    VALUES (?, ?, ?, ?, ?, ?, ?)",
                    sequences)
            else:
                cursor.executemany("INSERT INTO genome \
                    (T_NO, G_TAXID, G_SEQ_ID, G_SEQ_TYPE, G_DESCRIPTION, G_SEQUENCE, B_NO) \
                    VALUES (%s, %s, %s, %s, %s, %s, %s)",
                    sequences)
            self.db.commit()
    
        except Exception, e:
            QtGui.QMessageBox.critical(self, "Error", str(e[1]))
            self.db.rollback()
        
        finally:
            cursor.close()
            
        self.text.append("<br>" + str(reccount) + " record(s)  processed<br>")
    
    def habitats(self):
        with self.db:
            cursor = self.db.cursor()
            cursor.execute("SELECT COUNT(*) from habitats")
            n = int(cursor.fetchone()[0])
            if n > 0:
                reply = QtGui.QMessageBox.question(self, "Confirmation", 
                        "Delete data?\nThis operation cannot be undone!", 
                        QtGui.QMessageBox.Yes, QtGui.QMessageBox.No)
                if reply == QtGui.QMessageBox.Yes:
                        cursor.execute("DROP TABLE IF EXISTS habitats")
                else:
                    return
        
        SEARCH_BASE = 'http://apiv3.iucnredlist.org/api/v3/'
        SEARCH_HABITATS = "habitats/species/name/"
        TOKEN = "<YOUR_API_KEY>"
            
        reccount = 0    
        habitats = []
        RefNo = 0
        self.text.append("<br><b>Fetching habitat data from IUCN...</b><br>")
        for i in range(len(self.data)):
            #-- For each taxon in databaee
            TaxNo = i + 1
            Genus = self.data[i][5]
            Species = self.data[i][6]
            Subsp = self.data[i][8]
            Name = Genus + ' ' + Species + iif(len(Subsp) > 0, ' ' + Subsp, "")
            
            #-- Get conservation status from IUCN
            k = 0
            Place = ""
            url = SEARCH_BASE + SEARCH_HABITATS + Name + "?token=" + TOKEN
            results = simplejson.load(urllib.urlopen(url))
            for j in range(len(results)):
                result = results['result']
                try:
                    Habitat = result[j]['habitat']
                    k += 1
                    item = (TaxNo, Place, Habitat, RefNo)
                    habitats.append(item)
                except IndexError:
                    Habitat = None
                    if k < 1: k = 0
            self.text.append(str(reccount + 1) + " <i>" + Genus + ' ' + Species + ' ' + Subsp + "</i> " + ' - ' + str(k) + " habitat(s)")
                        
            reccount += 1
            self.text.repaint()
    
            # Sleep for two seconds to prevent IP blocking from IUCN
            time.sleep(2)
                
        #-- Insert records into the database 
        try:
            cursor = self.db.cursor()
            if self.adapter == "SQLite":
                cursor.executemany("INSERT INTO habitats \
                    (T_NO, H_PLACE, H_HABITAT, B_NO) \
                    VALUES (?, ?, ?, ?)",
                    habitats)
            else:
                cursor.executemany("INSERT INTO habitats \
                    (T_NO, H_PLACE, H_HABITAT, B_NO) \
                    VALUES (%s, %s, %s, %s)",
                    habitats)
            self.db.commit()
    
        except Exception, e:
            QtGui.QMessageBox.critical(self, "Error", str(e[1]))
            self.db.rollback()

        finally:
            cursor.close()
            
        self.text.append("<br>" + str(reccount) + " record(s)  processed<br>")
    
    def notes(self):
        import wikipedia
        
        with self.db:
            cursor = self.db.cursor()
            cursor.execute("SELECT COUNT(*) from notes")
            n = int(cursor.fetchone()[0])
            if n > 0:
                reply = QtGui.QMessageBox.question(self, "Confirmation", 
                        "Delete data?\nThis operation cannot be undone!", 
                        QtGui.QMessageBox.Yes, QtGui.QMessageBox.No)
                if reply == QtGui.QMessageBox.Yes:
                        cursor.execute("DROP TABLE IF EXISTS notes")
                else:
                    return
                    
        reccount = 0
        notes = []
        RefNo = 0
        self.text.append("<br><b>Fetching text snippets from Wikipedia...</b><br>")
        for i in range(len(self.data)):
            #-- For each taxon in databaee
            TaxNo = i + 1
            Genus = self.data[i][5]
            Species = self.data[i][6]
            Subsp = self.data[i][8]
            Name = Genus + ' ' + Species + ' ' + Subsp
    
            #-- Get a text snippet from Wikipedia
            try:
                Summary = wikipedia.summary(Name.replace(' ', '_'), sentences=1)
                item = (TaxNo, Summary.encode('latin-1', 'replace'), RefNo)
                notes.append(item)
            except:
                Summary = ""
            self.text.append(str(reccount + 1) + " <i>" + Genus + ' ' + Species + ' ' + Subsp + "</i>" + ' - ' + iif(len(Summary) > 0, str(1), str(0)) + " snippet(s)")
    
            reccount += 1
            self.text.repaint()
            
            # Sleep for one second to prevent IP blocking from Wikipedia
            time.sleep(1)
    
        #-- Insert records into the database 
        try:
            cursor = self.db.cursor()
            if self.adapter == "SQLite":
                cursor.executemany("INSERT INTO notes \
                    (T_NO, N_NOTE, B_NO) \
                    VALUES (?, ?, ?)",
                    notes)
            else:
                cursor.executemany("INSERT INTO notes \
                    (T_NO, N_NOTE, B_NO) \
                    VALUES (%s, %s, %s)",
                    notes)
            self.db.commit()
    
        except Exception, e:
            QtGui.QMessageBox.critical(self, "Error", str(e[1]))
            self.db.rollback()
        
        finally:
            cursor.close()
            
        self.text.append("<br>" + str(reccount) + " record(s)  processed<br>")
    
    def synonyms(self):
        with self.db:
            cursor = self.db.cursor()
            cursor.execute("SELECT COUNT(*) from synonyms")
            n = int(cursor.fetchone()[0])
            if n > 0:
                reply = QtGui.QMessageBox.question(self, "Confirmation", 
                        "Delete data?\nThis operation cannot be undone!", 
                        QtGui.QMessageBox.Yes, QtGui.QMessageBox.No)
                if reply == QtGui.QMessageBox.Yes:
                        cursor.execute("DROP TABLE IF EXISTS synonyms")
                else:
                    return
        
        reccount = 0
        synonymdata = []
        RefNo = 0
        self.text.append("<br><b>Fetching synonyms from CoL...</b><br>")
        for i in range(len(self.data)):
            #-- For each taxon in databaee
            TaxNo = i + 1
            Genus = self.data[i][5]
            Species = self.data[i][6]
            Subsp = self.data[i][8]
            Name = Genus + ' ' + Species + ' ' + Subsp
                
            #-- Retrieve data from CoL
            result = urllib.urlopen("http://www.catalogueoflife.org/col/webservice?name=" + urllib.quote_plus(Name) + "&response=full").read()
            root = ET.XML(result)
                
            #-- Get a list of synonyms
            synonyms = root.xpath("result/synonyms/synonym")
            self.text.append(str(reccount + 1) + " <i>" + Genus + ' ' + Species + ' ' + Subsp + "</i> " + ' - ' + str(len(synonyms)) + " synonym(s)")
                
            #-- Loop through the synonyms
            SGAuthor = None
            SSubgen = None
            SSAuthor = None
            SSubsp = None
            SRank = None
            for synonym in synonyms:
                SStatus = synonym.xpath("name_status/text()")[0]
                SGenus = synonym.xpath("genus/text()")[0] 
                SSpecies = synonym.xpath("species/text()")[0]
                try:
                    SAuthor = synonym.xpath("author/text()")[0]
                    SRank = synonym.xpath("infraspecies_marker/text()")[0]
                    SSubsp = synonym.xpath("infraspecies/text()")[0]
                except:
                    SAuthor = unicode_to_ascii(SAuthor)
                    pass
                item = (TaxNo, SStatus, SGenus, SGAuthor, SSubgen, SSpecies, SAuthor, SRank, SSubsp, SSAuthor, RefNo)
                synonymdata.append(item)
                    
            reccount += 1
            self.text.repaint()
            
            # Sleep for one second to prevent IP blocking from CoL
            time.sleep(1)
                    
        #-- Insert records into the database 
        try:
            cursor = self.db.cursor()
            if self.adapter == "SQLite":
                cursor.executemany("INSERT INTO synonyms \
                    (T_NO, S_STATUS, S_GENUS, S_G_AUTHOR, S_SUBGENUS, S_SPECIES, S_S_AUTHOR, S_RANK, S_SUBSP, S_SP_AUTHOR, B_NO) \
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", 
                    synonymdata)
            else:
                cursor.executemany("INSERT INTO synonyms \
                    (T_NO, S_STATUS, S_GENUS, S_G_AUTHOR, S_SUBGENUS, S_SPECIES, S_S_AUTHOR, S_RANK, S_SUBSP, S_SP_AUTHOR, B_NO) \
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)", 
                    synonymdata)
            self.db.commit()
                
        except Exception, e:
            QtGui.QMessageBox.critical(self, "Error", str(e[1]))
            self.db.rollback()
            pass
            
        finally:
            cursor.close()
            
        self.text.append("<br>" + str(reccount) + " record(s)  processed<br>")
    
    def commonnames(self):
        with self.db:
            cursor = self.db.cursor()
            cursor.execute("SELECT COUNT(*) from commonnames")
            n = int(cursor.fetchone()[0])
            if n > 0:
                reply = QtGui.QMessageBox.question(self, "Confirmation", 
                        "Delete data?\nThis operation cannot be undone!", 
                        QtGui.QMessageBox.Yes, QtGui.QMessageBox.No)
                if reply == QtGui.QMessageBox.Yes:
                        cursor.execute("DROP TABLE IF EXISTS commonnames")
                else:
                    return
        
        reccount = 0
        commonnamesdata = []
        RefNo = 0
        self.text.append("<br><b>Fetching common names from CoL...</b><br>")
        for i in range(len(self.data)):
            #-- For each taxon in databaee
            TaxNo = i + 1
            Genus = self.data[i][5]
            Species = self.data[i][6]
            Subsp = self.data[i][8]
            Name = Genus + ' ' + Species + ' ' + Subsp
                
            #-- Retrieve data from CoL
            result = urllib.urlopen("http://www.catalogueoflife.org/col/webservice?name=" + urllib.quote_plus(Name) + "&response=full").read()
            root = ET.XML(result)
                
            #-- Get a list of common names
            common_names = root.xpath("result/common_names/common_name")
            self.text.append(str(reccount + 1) + " <i>" + Genus + ' ' + Species + ' ' + Subsp + "</i> " + ' - ' + str(len(common_names)) + " common name(s)")
            
            #-- Loop through the common names
            VCountry = None
            VLang = None
            for common_name in common_names:
                VName = common_name.xpath("name/text()")[0]
                try:
                    VCountry = common_name.xpath("country/text()")[0]
                    VLang = common_name.xpath("language/text()")[0]
                except:
                    pass
                if is_ascii(VName):
                    item = (TaxNo, VName, VCountry, VLang, RefNo)
                    commonnamesdata.append(item)
                    
            reccount += 1
            self.text.repaint()
            
            # Sleep for one second to prevent IP blocking from CoL
            time.sleep(1)
                    
        #-- Insert records into the database 
        try:
            cursor = self.db.cursor()
            if self.adapter == "SQLite":
                cursor.executemany("INSERT INTO commonnames \
                    (T_NO, V_NAME, V_COUNTRY, V_LANGUAGE, B_NO) \
                    VALUES (?, ?, ?, ?, ?)",
                    commonnamesdata)
            else:
                cursor.executemany("INSERT INTO commonnames \
                    (T_NO, V_NAME, V_COUNTRY, V_LANGUAGE, B_NO) \
                    VALUES (%s, %s, %s, %s, %s)",
                    commonnamesdata)
            self.db.commit()
                
        except Exception, e:
            QtGui.QMessageBox.critical(self, "Error", str(e[1]))
            self.db.rollback()
            
        finally:
            cursor.close()
            
        self.text.append("<br>" + str(reccount) + " record(s)  processed<br>")
    
    def taxa(self):
        with self.db:
            cursor = self.db.cursor()
            cursor.execute("SELECT COUNT(*) from taxa")
            n = int(cursor.fetchone()[0])
            if n > 0:
                reply = QtGui.QMessageBox.question(self, "Confirmation", 
                        "Delete data?\nThis operation cannot be undone!", 
                        QtGui.QMessageBox.Yes, QtGui.QMessageBox.No)
                if reply == QtGui.QMessageBox.Yes:
                        cursor.execute("DROP TABLE IF EXISTS taxa")
                        cursor.execute("DROP TABLE IF EXISTS highertaxa")
                else:
                    return
        
        reccount = 0
        taxa = []
        highertaxa = []
        self.text.append("<br>Processing species checklist...<br>")
        for i in range(len(self.data)):
            #--- For each species in checklist spreadsheet
            TaxNo = i + 1
            Kingdom = self.data[i][0]
            Phylum = self.data[i][1]
            Class = self.data[i][2]
            Order = self.data[i][3]
            Family = self.data[i][4]
            Genus = self.data[i][5]
            Species = self.data[i][6]
            Author = self.data[i][7]
            Subsp = self.data[i][8]
            Status = self.data[i][9]
            self.text.append(str(reccount + 1) + " <i>" + Genus + ' ' + Species + ' ' + Subsp + "</i> ")
            
            GAuthor = None
            Subgenus = None
            if len(Subsp) == 0:
                Rank = "Species"
                SName = None
            else:
                if Subsp.find("var.") != -1 or Subsp.find("subsp.") != -1 or Subsp.find("ssp.") != -1:
                    Rank = Subsp.split('.')[0].capitalize() + '.'
                    SName = Subsp.split('.')[1].strip()
                else:
                    Rank = "Subsp."
                    SName = Subsp.strip()
                    
            SAuthor = None
            Subphylum = None
            Subclass = None
            Suborder = None
            Superfamily = None
            Subfamily = None
            Tribe = None
            RefNo = 0
        
            highertaxon = (TaxNo, Kingdom, Phylum, Subphylum, Class, Subclass, Order, Suborder, Family, Superfamily, Subfamily, Tribe)
            taxon = (TaxNo, Status, Genus, GAuthor, Subgenus, Species, Author, Rank, SName, SAuthor, RefNo)
            highertaxa.append(highertaxon)
            taxa.append(taxon)

            reccount += 1
            self.text.repaint()

        #-- Insert records into the database 
        try:
            cursor = self.db.cursor()
            if self.adapter == "SQLite":
                cursor.executemany("INSERT INTO taxa \
                    (T_NO, T_STATUS, T_GENUS, T_G_AUTHOR, T_SUBGENUS, T_SPECIES, T_S_AUTHOR, T_RANK, T_SUBSP, T_SP_AUTHOR, B_NO) \
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", 
                    taxa)
            else:
                cursor.executemany("INSERT INTO taxa \
                    (T_NO, T_STATUS, T_GENUS, T_G_AUTHOR, T_SUBGENUS, T_SPECIES, T_S_AUTHOR, T_RANK, T_SUBSP, T_SP_AUTHOR, B_NO) \
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)", 
                    taxa)
            self.db.commit()
        
            if self.adapter == "SQLite":
                cursor.executemany("INSERT INTO highertaxa \
                    (T_NO, T_KINGDOM, T_PHYLUM, T_SUBPHYLUM, T_CLASS, T_SUBCLASS, T_ORDER, T_SUBORDER, T_FAMILY, T_SUPERFAMILY, T_SUBFAMILY, T_TRIBE) \
                    VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", 
                    highertaxa)
            else:
                cursor.executemany("INSERT INTO highertaxa \
                    (T_NO, T_KINGDOM, T_PHYLUM, T_SUBPHYLUM, T_CLASS, T_SUBCLASS, T_ORDER, T_SUBORDER, T_FAMILY, T_SUPERFAMILY, T_SUBFAMILY, T_TRIBE) \
                    VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)", 
                    highertaxa)
            self.db.commit()
    
        except Exception, e:
            QtGui.QMessageBox.critical(self, "Error", str(e[1]))
            self.db.rollback()

        finally:
            cursor.close()
        
        self.text.append("<br>" + str(reccount) + " record(s)  processed<br>")
    
    def openFile(self):
        self.filename = QtGui.QFileDialog.getOpenFileName(self, "Open File", os.getenv("HOME"),
                    "Spreadsheets (*.csv *.ods *.xls *.xlsx)")
        if not self.filename.isEmpty():
            self.text.clear()
            self.text.setHtml(
                """Builds biodiversity databases from species checklists. 
                <br>&copy; 2016 Mauro J. Cavalcanti. 
                <br>Ecoinformatics Studio, Rio de Janeiro, Brazil. 
                <br>E-mail: maurobio@gmail.com""")
            self.readData(str(self.filename))
            self.connectDb()
            self.createTables()
            self.taxa()
        
    def about(self):
        QtGui.QMessageBox.about(self, "About Feronia",
                """<b>Feronia</b> v {0}
                <p>Copyright &copy;2016 Mauro J. Cavalcanti.
                <p>This application is an open-source, cross-platform
                desktop tool for building biodiversity databases by 
                harvesting data across several distributed sources 
                (CoL, EOL, GBIF, NCBI, Wikipedia, Google Scholar),
                on the basis of a user-provided species checklist.
                <p>Python {1} - Qt {2} - PyQt {3} on {4}""".format(
                __version__, platform.python_version(),
                QtCore.QT_VERSION_STR, QtCore.PYQT_VERSION_STR,
                platform.system()))
                
    def closeEvent(self, event):
        reply = QtGui.QMessageBox.question(self, "Confirmation", 
                        "Exit program?", 
                        QtGui.QMessageBox.Yes, QtGui.QMessageBox.No)
        if reply == QtGui.QMessageBox.Yes:
            f = open(os.path.splitext(str(self.filename))[0] + ".htm", 'a')
            filedata = self.text.toHtml()
            f.write(filedata)
            f.close()
            event.accept()
        else:
            event.ignore()
                
def main():
    app = QtGui.QApplication(sys.argv)
    main = MainWindow()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()